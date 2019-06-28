"""Main program's implementation."""
from shutil import rmtree
from datetime import datetime
import os
import openpyxl
import numpy as np
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from style import style_ws, style_range
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from helper import *

NAME = [
    "NFLX", "AAPL", "PG", "ADS", "AMGN", "AMZN", "CBRE", "COST", "CVX", "DAL", "FB",
    "GOOGL", "MMM", "NKE", "QCOM", "T", "TRIP", "NVIDIA"
]  # GS
IS = "Income Statement"
BS = "Balance Sheet"
CF = "Cashflow Statement"
YRS_TO_CONSIDER = 5

"""
TODO:
    Optimize searched label.
    All divide by zero errors.
    Red and blue style.
    Identation for style.
"""


class ValuationMachine:
    def __init__(self, name, growth_rates):
        self.name = name
        self.growth_rates = growth_rates
        self.yrs_to_predict = len(growth_rates)
        self.is_df, self.bs_df, self.cf_df, self.mkt_df = None, None, None, None
        self.is_unit, self.bs_unit, self.cf_unit, self.mkt_unit = None, None, None, None
        self.fye, self.wb, self.case_col, self.case_rate_cells = None, None, None, []

    def read(self):
        self.is_df = pd.read_excel("asset/{} IS.xlsx".format(self.name), header=4, index_col=0)
        self.bs_df = pd.read_excel("asset/{} BS.xlsx".format(self.name), header=4, index_col=0)
        self.cf_df = pd.read_excel("asset/{} CF.xlsx".format(self.name), header=4, index_col=0)
        self.mkt_df = pd.read_excel("asset/{} MKT.xlsx".format(self.name), index_col=0)

    def preprocess(self):
        self.fye = self.is_df.columns[1]
        self.is_df = preprocess(self.is_df, YRS_TO_CONSIDER)
        self.bs_df = preprocess(self.bs_df, YRS_TO_CONSIDER)
        self.cf_df = preprocess(self.cf_df, YRS_TO_CONSIDER)

    def get_units(self):
        self.is_unit, self.bs_unit = get_unit(self.is_df), get_unit(self.bs_df)
        self.cf_unit, self.mkt_unit = get_unit(self.cf_df), get_unit(self.mkt_df)

        # FIXME
        if self.is_unit != self.bs_unit or self.bs_unit != self.cf_unit:
            print("{} is:{}, bs:{}, cf:{}".format(
                self.name, self.is_unit, self.bs_unit, self.cf_unit
            ))

        if self.mkt_unit != self.is_unit:
            if self.mkt_unit == 'm':
                self.mkt_multiplier = 0.001
            else:
                print("Market unit bigger")
                exit(1)
        else:
            self.mkt_multiplier = 1

        if self.is_unit != self.bs_unit:  # assume CF and IS have the same units
            if self.bs_unit == 'b':
                self.extra_bs = "/1000"
                self.extra_cf = "*1000"
            else:
                self.extra_bs = "*1000"
                self.extra_cf = "/1000"
        else:
            self.extra_bs = ""
            self.extra_cf = ""

    def slice_data(self):
        if searched_label(self.is_df.index, "EBITDA"):
            is_search = self.is_df.index.get_loc(searched_label(self.is_df.index, "EBITDA"))
        else:
            is_search = self.is_df.index.get_loc(
                searched_label(self.is_df.index, "dividends per share")
            )
        if isinstance(is_search, int):
            self.is_df = self.is_df[:is_search + 1]
        else:
            self.is_df = self.is_df[:np.where(is_search)[-1][-1] + 1]
        bs_search = self.bs_df.index.get_loc(searched_label(
            self.bs_df.index, "total liabilities & shareholder equity")
        )
        if isinstance(bs_search, int):
            self.bs_df = self.bs_df[:bs_search + 1]
        else:
            self.bs_df = self.bs_df[:np.where(bs_search)[-1][-1] + 1]
        cf_search = self.cf_df.index.get_loc(
            searched_label(self.cf_df.index, "net change in cash")
        )
        if isinstance(cf_search, int):
            self.cf_df = self.cf_df[:cf_search + 1]
        else:
            self.cf_df = self.cf_df[:np.where(cf_search)[-1][-1] + 1]

    def style(self):
        # Stylize excel sheets and output excel
        self.wb = openpyxl.Workbook()
        self.wb['Sheet'].title = "Summary"
        self.wb.create_sheet(IS)
        self.wb.create_sheet(BS)
        self.wb.create_sheet(CF)
        for r in dataframe_to_rows(self.is_df):
            self.wb[IS].append(r)
        for r in dataframe_to_rows(self.bs_df):
           self.wb[BS].append(r)
        for r in dataframe_to_rows(self.cf_df):
            self.wb[CF].append(r)
        style_ws(self.wb[IS], IS, self.is_df, self.bs_df, self.cf_df, self.fye, self.is_unit)
        style_ws(self.wb[BS], BS, self.is_df, self.bs_df, self.cf_df, self.fye, self.bs_unit)
        style_ws(self.wb[CF], CF, self.is_df, self.bs_df, self.cf_df, self.fye, self.cf_unit)

    def process_is(self):
        """Manipulates income statement."""
        # Short-hands
        yrs_to_predict = self.yrs_to_predict
        last_given_yr = self.is_df.columns[-1]

        # Diluted shares outstanding
        self.is_df.loc["Diluted Shares Outstanding"] = np.nan
        self.is_df.loc["Diluted Shares Outstanding"].iloc[-1] = self.mkt_df.loc[
            searched_label(self.mkt_df.index, "fully diluted equity capitalization")
        ][0] * self.mkt_multiplier

        # Declare income statement labels
        sales = searched_label(self.is_df.index, "total sales")
        cogs = searched_label(self.is_df.index, "cost of goods sold COGS including")
        self.is_df.index = [  # change label name from incl. to excl.
            i if i != cogs else "Cost of Goods Sold (COGS) excl. D&A" \
            for i in list(self.is_df.index)
        ]
        cogs = "Cost of Goods Sold (COGS) excl. D&A"
        gross_income = searched_label(self.is_df.index, "gross income")
        sgna = searched_label(self.is_df.index, "sg&a expense")
        other_expense = searched_label(self.is_df.index, "other expense")
        ebit = searched_label(self.is_df.index, "ebit operating income")
        nonoperating_income = searched_label(self.is_df.index, "nonoperating income net")
        interest_expense = searched_label(self.is_df.index, "interest expense")
        unusual = searched_label(self.is_df.index, "unusual expense")
        income_tax = searched_label(self.is_df.index, "income taxes")
        diluted_eps = searched_label(self.is_df.index, "eps diluted")
        net_income = searched_label(self.is_df.index, "net income")
        div_per_share = searched_label(self.is_df.index, "dividends per share")
        ebitda = searched_label(self.is_df.index, "ebitda")
        diluted_share_outstanding = "Diluted Shares Outstanding"

        # Drop EBITDA if exists
        if ebitda:
            self.is_df.drop(ebitda, inplace=True)

        # Insert pretax income row before income taxes
        pretax_df = pd.DataFrame(
            {
                yr: '={}+{}-{}-{}'.format(
                    excel_cell(self.is_df, ebit, yr),
                    excel_cell(self.is_df, nonoperating_income, yr),
                    excel_cell(self.is_df, interest_expense, yr),
                    excel_cell(self.is_df, unusual, yr)
                ) for yr in self.is_df.columns
            }, index=["Pretax Income"]
        )
        self.is_df = insert_before(self.is_df, pretax_df, income_tax)
        pretax = "Pretax Income"

        # Insert depreciation & amortization expense before SG&A expense
        dna_expense_df = pd.DataFrame(
            {               
                yr: "='{}'!".format(CF) + excel_cell(
                    self.cf_df, searched_label(
                        self.cf_df.index, "depreciation depletion & amortization expense"
                    ), yr
                ) for yr in self.is_df.columns
            }, index=["Depreciation & Amortization Expense"]
        )
        self.is_df = insert_before(self.is_df, dna_expense_df, sgna)
        dna_expense = "Depreciation & Amortization Expense"

        # Write formulas for COGS excl. D&A
        self.is_df.loc[cogs] = [
            '={}-{}'.format(self.is_df.at[cogs, yr], excel_cell(self.is_df, dna_expense, yr))
            for yr in self.is_df.columns
        ]

        # Add driver rows to income statement
        add_empty_row(self.is_df)
        self.is_df.loc["Driver Ratios"] = np.nan
        add_growth_rate_row(self.is_df, sales, "Sales Growth %")
        sales_growth = "Sales Growth %"
        add_empty_row(self.is_df)
        initialize_ratio_row(self.is_df, cogs, sales, "COGS Sales Ratio")
        cogs_ratio = "COGS Sales Ratio"
        add_empty_row(self.is_df)
        initialize_ratio_row(self.is_df, sgna, sales, "SG&A Sales Ratio")
        sgna_ratio = "SG&A Sales Ratio"
        add_empty_row(self.is_df)
        initialize_ratio_row(self.is_df, dna_expense, sales, "D&A Sales Ratio")
        dna_ratio = "D&A Sales Ratio"
        add_empty_row(self.is_df)
        initialize_ratio_row(self.is_df, unusual, ebit, "Unusual Expense EBIT Ratio")
        unusual_ratio = "Unusual Expense EBIT Ratio"
        add_empty_row(self.is_df)
        initialize_ratio_row(self.is_df, income_tax, pretax, "Effective Tax Rate")
        effective_tax = "Effective Tax Rate"

        # Add prediction years
        for i in range(yrs_to_predict):
            add_yr_column(self.is_df)

        # Set excel cell of case
        self.case_col = chr(ord('B') + len(self.is_df.columns))

        # Case insertions
        def add_case(label, ratio_input=None):
            # Label has to be exactly correct
            prev_letter = chr(ord('B') + len(self.is_df.columns) - yrs_to_predict)
            start_num = self.is_df.index.get_loc(label) + 5
            case_df = pd.DataFrame(
                {yr: [np.nan] * 5 for yr in self.is_df.columns},
                index=["Bull", "Upside", "Base", "Downside", "Bear"]
            )
            case_rate_cell = '${}${}'.format(chr(ord(prev_letter) + 6), start_num + 2)
            for i, yr in enumerate(self.is_df.columns[-yrs_to_predict:]):
                cur_letter = chr(ord(prev_letter) + i + 1)
                self.is_df.at[label, yr] = '=CHOOSE(${1}$2, {0}{2}, {0}{3}, {0}{4}, {0}{5}, {0}{6})'.format(
                    cur_letter, self.case_col, start_num, start_num + 1, start_num + 2, start_num + 3, start_num + 4
                )
                if ratio_input:
                    temp = ratio_input[i]
                else:
                    temp = '={}{}'.format(prev_letter, start_num - 1)
                case_df[yr] = [
                    '={}{}+{}'.format(cur_letter, start_num + 1, case_rate_cell),
                    '={}{}+{}'.format(cur_letter, start_num + 2, case_rate_cell), temp,
                    '={}{}-{}'.format(cur_letter, start_num + 2, case_rate_cell),
                    '={}{}-{}'.format(cur_letter, start_num + 3, case_rate_cell)
                ]
            self.is_df = insert_after(self.is_df, case_df, label)
            self.case_rate_cells.append(case_rate_cell[1] + case_rate_cell[3:])
        add_case(sales_growth, self.growth_rates)
        add_case(cogs_ratio)
        add_case(sgna_ratio)

        # Write formulas for EBITDA
        ebitda_df = pd.DataFrame(
            {
                yr: '={}+{}'.format(
                    excel_cell(self.is_df, dna_expense, yr), excel_cell(self.is_df, ebit, yr)
                ) for yr in self.is_df.columns
            }, index=[ebitda]
        )
        self.is_df = insert_after(self.is_df, ebitda_df, div_per_share)

        # Write formulas for driver ratios
        initialize_ratio_row(self.is_df, div_per_share, diluted_eps, "Dividend Payout Ratio")
        initialize_ratio_row(self.is_df, ebitda, sales, "EBITDA Margin", sales_growth)
        self.is_df.loc[dna_ratio].iloc[-yrs_to_predict:] = self.is_df.loc[dna_ratio, last_given_yr]
        driver_extend(self.is_df, unusual_ratio, "avg", last_given_yr, yrs_to_predict)
        driver_extend(self.is_df, diluted_share_outstanding, "avg",
                      last_given_yr, yrs_to_predict, 3)
        driver_extend(self.is_df, effective_tax, "avg", last_given_yr, yrs_to_predict)

        # Calculate fixed variables
        fixed_extend(self.is_df, other_expense, "prev", yrs_to_predict)
        fixed_extend(self.is_df, nonoperating_income, "prev", yrs_to_predict)
        fixed_extend(self.is_df, interest_expense, "prev", yrs_to_predict)
        fixed_extend(self.is_df, div_per_share, "prev", yrs_to_predict)  # FIXME
        fixed_extend(self.is_df, diluted_share_outstanding, "prev", yrs_to_predict)

        # Write formula for net income
        self.is_df.loc[net_income] = [
            '={}-{}'.format(
                excel_cell(self.is_df, pretax, yr),
                excel_cell(self.is_df, income_tax, yr)
            ) for yr in self.is_df.columns
        ]

        for i in range(yrs_to_predict):
            cur_yr = self.is_df.columns[-yrs_to_predict + i]
            prev_yr = self.is_df.columns[-yrs_to_predict + i - 1]

            # Write formulas
            self.is_df.at[sales, cur_yr] = '={}*(1+{})'.format(
                excel_cell(self.is_df, sales, prev_yr),
                excel_cell(self.is_df, sales_growth, cur_yr)
            )
            self.is_df.at[cogs, cur_yr] = '={}*{}'.format(
                excel_cell(self.is_df, sales, cur_yr), excel_cell(self.is_df, cogs_ratio, cur_yr)
            )
            self.is_df.at[gross_income, cur_yr] = '={}-{}'.format(
                excel_cell(self.is_df, sales, cur_yr), excel_cell(self.is_df, cogs, cur_yr)
            )
            self.is_df.at[dna_expense, cur_yr] = '={}*{}'.format(
                excel_cell(self.is_df, sales, cur_yr), excel_cell(self.is_df, dna_ratio, cur_yr)
            )
            self.is_df.at[sgna, cur_yr] = '={}*{}'.format(
                excel_cell(self.is_df, sales, cur_yr), excel_cell(self.is_df, sgna_ratio, cur_yr)
            )
            self.is_df.at[ebit, cur_yr] = '={}-{}'.format(
                excel_cell(self.is_df, gross_income, cur_yr),
                sum_formula(self.is_df, ebit, cur_yr, gross_income, 1)
            )
            self.is_df.at[unusual, cur_yr] = '={}*{}'.format(
                excel_cell(self.is_df, ebit, cur_yr),
                excel_cell(self.is_df, unusual_ratio, cur_yr)
            )
            self.is_df.at[pretax, cur_yr] = '={}+{}-{}-{}'.format(
                excel_cell(self.is_df, ebit, cur_yr),
                excel_cell(self.is_df, nonoperating_income, cur_yr),
                excel_cell(self.is_df, interest_expense, cur_yr),
                excel_cell(self.is_df, unusual, cur_yr)
            )
            self.is_df.at[income_tax, cur_yr] = '={}*{}'.format(
                 excel_cell(self.is_df, pretax, cur_yr),
                 excel_cell(self.is_df, effective_tax, cur_yr)
            )
            self.is_df.at[diluted_eps, cur_yr] = '={}/{}'.format(
                excel_cell(self.is_df, net_income, cur_yr),
                excel_cell(self.is_df, diluted_share_outstanding, cur_yr)
            )

            if isinstance(self.is_df.at[ebitda, cur_yr], str):  # two rows with label EBITDA
                self.is_df.at[ebitda, cur_yr] = '={}+{}'.format(
                    excel_cell(self.is_df, dna_expense, cur_yr),
                    excel_cell(self.is_df, ebit, cur_yr)
                )
            else:
                self.is_df.at[ebitda, cur_yr] = [
                    np.nan,
                ]
        empty_unmodified(self.is_df, yrs_to_predict)


    def process_bs(self):
        """Manipulates balance sheet."""
        # Short-hands
        yrs_to_predict = self.yrs_to_predict
        last_given_yr = self.bs_df.columns[-1]

        # Balance sheet labels
        st_receivables = searched_label(self.bs_df.index, "short term receivables")
        cash_st_investments = searched_label(self.bs_df.index, "cash short term investments")
        inventories = searched_label(self.bs_df.index, "inventories")
        other_cur_assets = searched_label(self.bs_df.index, "other current asset")
        total_cur_assets = searched_label(self.bs_df.index, "total current asset")
        net_property_plant_equipment = searched_label(self.bs_df.index,
                                                      "net property plant equipment")
        total_assets = searched_label(self.bs_df.index, "total assets")
        accounts_payable = searched_label(self.bs_df.index, "accounts payable")
        other_cur_liabilities = searched_label(self.bs_df.index, "other current liabilities")
        total_cur_liabilities = searched_label(self.bs_df.index, "total current liabilities")
        total_liabilities = searched_label(self.bs_df.index, "total liabilities")
        total_equity = searched_label(self.bs_df.index, "total equity")
        total_liabilities_n_shareholders_equity = searched_label(
            self.bs_df.index, "total liabilities shareholders equity"
        )

        # Income statement labels
        sales = searched_label(self.is_df.index, "total sales")
        cogs = searched_label(self.is_df.index, "cost of goods sold cogs excl")
        net_income = searched_label(self.is_df.index, "net income")

        # Cash flow statement labels
        deprec_deplet_n_amort = searched_label(self.cf_df.index,
                                               "depreciation depletion amortization expense")
        capital_expenditures = searched_label(self.cf_df.index, "capital expenditures")
        cash_div_paid = searched_label(self.cf_df.index, "cash dividends paid", True)
        change_in_capital_stock = searched_label(self.cf_df.index, "change in capital stock")

        # Add driver rows to balance sheet
        add_empty_row(self.bs_df)
        self.bs_df.loc["Driver Ratios"] = np.nan
        # DSO
        if st_receivables:
            self.bs_df.loc["DSO"] = [
                "={}/'{}'!{}{}*365".format(
                    excel_cell(self.bs_df, st_receivables, yr), IS,
                    excel_cell(self.is_df, sales, yr), self.extra_cf
                ) for yr in self.bs_df.columns
            ]
            dso = "DSO"
        # Other current assets growth %
        if sum(self.bs_df.loc[other_cur_assets]) != 0:
            add_growth_rate_row(self.bs_df, other_cur_assets, "Other Current Assets Growth %")
            other_cur_assets_growth = "Other Current Assets Growth %"
        else:
            other_cur_assets_growth = None
        # DPO
        add_empty_row(self.bs_df)
        self.bs_df.loc["DPO"] = [
            "={}/'{}'!{}{}*366".format(
                excel_cell(self.bs_df, accounts_payable, yr), IS,
                excel_cell(self.is_df, cogs, yr), self.extra_cf
            ) for yr in self.bs_df.columns
        ]
        dpo = "DPO"
        # Miscellaneous Current Liabilities Growth %
        add_growth_rate_row(self.bs_df, other_cur_liabilities,
                            "Miscellaneous Current Liabilities Growth %")
        misc_cur_liabilities_growth = "Miscellaneous Current Liabilities Growth %"

        # Inventory turnober ratio
        if inventories:
            add_empty_row(self.bs_df)
            self.bs_df.loc["Inventory Turnover Ratio"] = np.nan
            self.bs_df.loc["Inventory Turnover Ratio"].iloc[1:] = [
                "='{}'!{}{}/({}+{})*2".format(
                    IS, excel_cell(self.is_df, cogs, self.bs_df.columns[i + 1]), self.extra_bs,
                    excel_cell(self.bs_df, inventories, self.bs_df.columns[i]),
                    excel_cell(self.bs_df, inventories, self.bs_df.columns[i+1])
                ) for i in range(len(self.bs_df.columns) - 1)
            ]
        inventory_ratio = "Inventory Turnover Ratio"

        # Add driver rows to cash flow statement
        add_empty_row(self.cf_df)
        self.cf_df.loc["Driver Ratios"] = np.nan
        # Capital Expenditure Revenue Ratio
        self.cf_df.loc["Capital Expenditure Revenue Ratio"] = [
            "=-{}/'{}'!{}".format(
                excel_cell(self.cf_df, capital_expenditures, yr), IS,
                excel_cell(self.is_df, sales, yr)
            ) for yr in self.cf_df.columns
        ]
        # Other Funds Net Operating CF Ratio
        net_operating_cf = searched_label(self.cf_df.index, "net operat cash flow cf")
        initialize_ratio_row(
            self.cf_df, searched_label(self.cf_df.index, "other funds"),
            net_operating_cf, "Other Funds Net Operating CF Ratio", net_operating_cf
        )

        # Add prediction years
        for i in range(yrs_to_predict):
            add_yr_column(self.bs_df)
        for i in range(yrs_to_predict):
            add_yr_column(self.cf_df)

        # Insert cash balance
        cash_balance_df = pd.DataFrame({yr: np.nan for yr in self.cf_df.columns},
                                       index=["Cash Balance"])
        self.cf_df = insert_after(self.cf_df, cash_balance_df, "net change in tax")
        cash_balance = searched_label(self.cf_df.index, "cash balance")

        # Inesrt working capital row
        wk_df = pd.DataFrame(
            {
                yr: ['={}-{}'.format(
                    excel_cell(self.bs_df, total_cur_assets, yr),
                    excel_cell(self.bs_df, total_cur_liabilities, yr)
                ), np.nan] for yr in self.bs_df.columns
            }, index=["Working Capital", np.nan]
        )
        self.bs_df = insert_before(self.bs_df, wk_df, "driver ratios")

        # Inesrt balance row
        balance_df = pd.DataFrame(
            {
                yr: ['={}-{}'.format(
                    excel_cell(self.bs_df, total_assets, yr),
                    excel_cell(self.bs_df, total_liabilities_n_shareholders_equity, yr)
                ), np.nan] for yr in self.bs_df.columns
            }, index=["Balance", np.nan]
        )
        self.bs_df = insert_before(self.bs_df, balance_df, "working capital")

        # Calculate driver ratios
        if st_receivables:
            self.bs_df.loc[dso].iloc[-yrs_to_predict:] = '=' + excel_cell(
                self.bs_df, dso, self.bs_df.columns[-yrs_to_predict - 2]
            )
        driver_extend(self.bs_df, dpo, "avg", last_given_yr, yrs_to_predict)
        driver_extend(self.bs_df, other_cur_assets_growth, "avg", last_given_yr, yrs_to_predict)
        driver_extend(self.bs_df, misc_cur_liabilities_growth, "avg",
                      last_given_yr, yrs_to_predict)
        driver_extend(self.bs_df, inventory_ratio, "avg", last_given_yr, yrs_to_predict)

        # Calculate total liabilities & shareholders' equity
        self.bs_df.loc[total_liabilities_n_shareholders_equity] = [
            '={}+{}'.format(
                excel_cell(self.bs_df, total_liabilities, yr),
                excel_cell(self.bs_df, total_equity, yr)
            ) for yr in self.bs_df.columns
        ]

        for i in range(yrs_to_predict):
            cur_yr = self.bs_df.columns[-yrs_to_predict + i]
            prev_yr = self.bs_df.columns[-yrs_to_predict + i - 1]

            # Calculate variables
            self.bs_df.at[cash_st_investments, cur_yr] = "='{}'!{}{}".format(
                CF, excel_cell(self.cf_df, cash_balance, cur_yr), self.extra_bs
            )
            if st_receivables:
                self.bs_df.at[st_receivables, cur_yr] = "={}/365*'{}'!{}{}".format(
                    excel_cell(self.bs_df, dso, cur_yr), IS, excel_cell(self.is_df, sales, cur_yr),
                    self.extra_bs
                )
            if other_cur_assets_growth:
                self.bs_df.at[other_cur_assets, cur_yr] = '={}*(1+{})'.format(
                    excel_cell(self.bs_df, other_cur_assets, prev_yr),
                    excel_cell(self.bs_df, other_cur_assets_growth, cur_yr)
                )
            self.bs_df.at[net_property_plant_equipment, cur_yr] = "={}-'{}'!{}{}-'{}'!{}{}".format(
                excel_cell(self.bs_df, net_property_plant_equipment, prev_yr), CF,
                excel_cell(self.cf_df, deprec_deplet_n_amort, cur_yr), self.extra_bs, CF,
                excel_cell(self.cf_df, capital_expenditures, cur_yr), self.extra_bs
            )
            self.bs_df.at[accounts_payable, cur_yr] = "={}/365*'{}'!{}{}".format(
                excel_cell(self.bs_df, dpo, cur_yr), IS, excel_cell(self.is_df, cogs, cur_yr),
                self.extra_bs
            )
            self.bs_df.at[other_cur_liabilities, cur_yr] = "={}*(1+{})".format(
                excel_cell(self.bs_df, other_cur_liabilities, prev_yr),
                excel_cell(self.bs_df, misc_cur_liabilities_growth, cur_yr)
            )
            self.bs_df.at[total_equity, cur_yr] = "={}+'{}'!{}{}+'{}'!{}{}".format(
                excel_cell(self.bs_df, total_equity, prev_yr), CF,
                excel_cell(self.cf_df, change_in_capital_stock, cur_yr), self.extra_bs, IS,
                excel_cell(self.is_df, net_income, cur_yr), self.extra_bs
            )
            if cash_div_paid:
                self.bs_df.at[total_equity, cur_yr] += "+'{}'!{}{}".format(
                    CF, excel_cell(self.cf_df, cash_div_paid, cur_yr), self.extra_bs
                )

            # Sums
            self.bs_df.at[total_cur_assets, cur_yr] = '=' + sum_formula(
                self.bs_df, total_cur_assets, cur_yr
            )
            self.bs_df.at[total_assets, cur_yr] = '={}+{}'.format(
                excel_cell(self.bs_df, total_cur_assets, cur_yr),
                sum_formula(self.bs_df, total_assets, cur_yr)
            )
            self.bs_df.at[total_cur_liabilities, cur_yr] = '=' + sum_formula(
                self.bs_df, total_cur_liabilities, cur_yr
            )
            self.bs_df.at[total_liabilities, cur_yr] = '={}+{}'.format(
                excel_cell(self.bs_df, total_cur_liabilities, cur_yr),
                sum_formula(self.bs_df, total_liabilities, cur_yr)
            )

        for label in self.bs_df.index[:self.bs_df.index.get_loc(total_liabilities)]:
            if pd.notna(label) and pd.notna(self.bs_df.loc[label].iloc[0]) \
                               and self.bs_df.loc[label].iloc[-1] == '0':
                fixed_extend(self.bs_df, label, 'prev', yrs_to_predict)

        empty_unmodified(self.bs_df, yrs_to_predict)

    def process_cf(self):
        """Manipulates cash flow statement."""
        # Short-hands
        yrs_to_predict = self.yrs_to_predict
        last_given_yr = self.cf_df.columns[-yrs_to_predict-1]

        # Cash flow statement labels
        net_income_cf = searched_label(self.cf_df.index, "net income starting line")
        deprec_deplet_n_amort = searched_label(self.cf_df.index,
                                               "depreciation depletion amortization")
        deferred_taxes = searched_label(self.cf_df.index,
                                        "deferred taxes & investment tax credit")
        other_funds = searched_label(self.cf_df.index, "other funds")
        funds_from_operations = searched_label(self.cf_df.index, "funds from operations")
        changes_in_working_capital = searched_label(self.cf_df.index,
                                                    "changes in working capital")
        net_operating_cf = searched_label(self.cf_df.index, "net operating cash flow")
        capital_expenditures = searched_label(self.cf_df.index, "capital expenditures")
        net_asset_acquisition = searched_label(self.cf_df.index, "net assets from acquisiton")
        fixed_assets_n_businesses_sale = searched_label(self.cf_df.index,
                                                        "fixed assets & of sale businesses")
        purchase_sale_of_investments = searched_label(self.cf_df.index,
                                                      "purchasesale of investments")
        net_investing_cf = searched_label(self.cf_df.index, "net investing cash flow")
        cash_div_paid = searched_label(self.cf_df.index, "cash dividends paid", precise=True)
        change_in_capital_stock = searched_label(self.cf_df.index, "change in capital stock")
        net_inssuance_reduction_of_debt = searched_label(self.cf_df.index,
                                                         "net issuance reduction of debt")
        net_financing_cf = searched_label(self.cf_df.index, "net financing cash flow")
        net_change_in_cash = searched_label(self.cf_df.index, "net change in cash")
        capex_ratio = "Capital Expenditure Revenue Ratio"
        other_funds_net_operating_ratio = "Other Funds Net Operating CF Ratio"

        # Income statement labels
        sales = searched_label(self.is_df.index, "total sales")
        deprec_amort_expense = searched_label(self.is_df.index,
                                              "depreciation amortization expense")
        net_income_is = searched_label(self.is_df.index, "net income")
        diluted_share_outstanding = searched_label(self.is_df.index, "diluted shares outstanding")
        div_per_share = searched_label(self.is_df.index, "dividends per share")

        # Balance sheet labels
        other_cur_assets = searched_label(self.bs_df.index, "other current assets")
        other_cur_liabilities = searched_label(self.bs_df.index, "other current liabilities")
        cash_st_investments = searched_label(self.bs_df.index, "cash short term investments")
        st_receivables = searched_label(self.bs_df.index, "short term receivables")
        accounts_payable = searched_label(self.bs_df.index, "accounts payable")
        lt_debt = searched_label(self.bs_df.index, "long term debt")

        # Insert cash balance
        self.cf_df.loc["Cash Balance"].iloc[-yrs_to_predict - 1:] = [
            "='{}'!{}{}".format(
                BS, excel_cell(self.bs_df, cash_st_investments, last_given_yr), self.extra_cf
            )
        ] + [
            '={}+{}'.format(
                excel_cell(
                    self.cf_df, "Cash Balance", self.cf_df.columns[-yrs_to_predict + i - 1]
                ),
                excel_cell(
                    self.cf_df, net_change_in_cash, self.cf_df.columns[-yrs_to_predict + i]
                )
            ) for i in range(yrs_to_predict)
        ]

        # Add levered free CF row
        self.cf_df.loc["Levered Free Cash Flow"] = [
            '={}+{}'.format(
                excel_cell(self.cf_df, net_operating_cf, yr),
                excel_cell(self.cf_df, capital_expenditures, yr)
            ) for yr in self.cf_df.columns
        ]
        levered_free_cf = "Levered Free Cash Flow"

        # Add levered free CF row growth %
        add_growth_rate_row(self.cf_df, levered_free_cf, "Levered Free Cash Flow Growth %")
        levered_free_cf_growth = "Levered Free Cash Flow Growth %"

        # Calculate driver ratios
        driver_extend(self.cf_df, capex_ratio, "avg", last_given_yr, yrs_to_predict)
        driver_extend(self.cf_df, other_funds_net_operating_ratio, "avg",
                      last_given_yr, yrs_to_predict)

        # Calculate fixed variables
        fixed_extend(self.cf_df, deferred_taxes, "zero", yrs_to_predict)
        fixed_extend(self.cf_df, other_funds, "zero", yrs_to_predict)
        fixed_extend(self.cf_df, net_asset_acquisition, "zero", yrs_to_predict)
        fixed_extend(self.cf_df, fixed_assets_n_businesses_sale, "zero", yrs_to_predict)
        fixed_extend(self.cf_df, purchase_sale_of_investments, "zero", yrs_to_predict)
        fixed_extend(self.cf_df, change_in_capital_stock, "prev", yrs_to_predict)

        # Calculate net operating CF
        self.cf_df.loc[net_operating_cf] = [
            '={}+{}'.format(
                excel_cell(self.cf_df, funds_from_operations, yr),
                excel_cell(self.cf_df, changes_in_working_capital, yr)
            ) for yr in self.cf_df.columns
        ]

        # Calculate net investing CF
        self.cf_df.loc[net_investing_cf] = [
            '=' + sum_formula(self.cf_df, net_investing_cf, yr) for yr in self.cf_df.columns
        ]

        # Calcualate net financing CF
        self.cf_df.loc[net_financing_cf] = [
            '=' + sum_formula(self.cf_df, net_financing_cf, yr) for yr in self.cf_df.columns
        ]

        # Calculate net change in cash
        self.cf_df.loc[net_change_in_cash] = [
            '={}+{}+{}'.format(
                excel_cell(self.cf_df, net_operating_cf, yr),
                excel_cell(self.cf_df, net_investing_cf, yr),
                excel_cell(self.cf_df, net_financing_cf, yr)
            ) for yr in self.cf_df.columns
        ]

        for i in range(yrs_to_predict):
            cur_yr = self.is_df.columns[-yrs_to_predict + i]
            prev_yr = self.is_df.columns[-yrs_to_predict + i - 1]

            # Calculate variables
            self.cf_df.at[net_income_cf, cur_yr] = "='{}'!{}".format(
                IS, excel_cell(self.is_df, net_income_is, cur_yr)
            )
            self.cf_df.at[deprec_deplet_n_amort, cur_yr] = "='{}'!{}".format(
                IS, excel_cell(self.is_df, deprec_amort_expense, cur_yr)
            )

            self.cf_df.at[funds_from_operations, cur_yr] = '=' + sum_formula(
                self.cf_df, funds_from_operations, cur_yr
            )

            self.cf_df.at[changes_in_working_capital, cur_yr] = "=SUM('{}'!{}:{}){}".format(
                BS, excel_cell(
                    self.bs_df,
                    self.bs_df.index[self.bs_df.index.get_loc(cash_st_investments) + 1],
                    prev_yr
                ), excel_cell(self.bs_df, other_cur_assets, prev_yr), self.extra_cf
            )
            self.cf_df.at[changes_in_working_capital, cur_yr] += "-SUM('{}'!{}:{}){}".format(
                BS, excel_cell(
                    self.bs_df,
                    self.bs_df.index[self.bs_df.index.get_loc(cash_st_investments) + 1],
                    cur_yr
                ), excel_cell(self.bs_df, other_cur_assets, cur_yr), self.extra_cf
            )
            self.cf_df.at[changes_in_working_capital, cur_yr] += "+SUM('{}'!{}:{}){}".format(
                BS, excel_cell(self.bs_df, accounts_payable, cur_yr),
                excel_cell(self.bs_df, other_cur_liabilities, cur_yr), self.extra_cf
            )
            self.cf_df.at[changes_in_working_capital, cur_yr] += "-SUM('{}'!{}:{}){}".format(
                BS, excel_cell(self.bs_df, accounts_payable, prev_yr),
                excel_cell(self.bs_df, other_cur_liabilities, prev_yr), self.extra_cf
            )

            self.cf_df.at[capital_expenditures, cur_yr] = "=-'{}'!{}*{}".format(
                IS, excel_cell(self.is_df, sales, cur_yr),
                excel_cell(self.cf_df, capex_ratio, cur_yr)
            )
            if cash_div_paid:
                self.cf_df.at[cash_div_paid, cur_yr] = "=-'{}'!{}*'{}'!{}".format(
                    IS, excel_cell(self.is_df, diluted_share_outstanding, cur_yr),
                    IS, excel_cell(self.is_df, div_per_share, cur_yr)
                )
            self.cf_df.at[net_inssuance_reduction_of_debt, cur_yr] = "='{}'!{}{}-'{}'!{}{}".format(
                BS, excel_cell(self.bs_df, lt_debt, cur_yr), self.extra_cf,
                BS, excel_cell(self.bs_df, lt_debt, prev_yr), self.extra_cf
            )
        empty_unmodified(self.cf_df, yrs_to_predict)

    def add_summary(self):
        """Note that number of years is fixed here."""
        years = self.is_df.columns[-self.yrs_to_predict - 4:-self.yrs_to_predict + 3]

        ws = self.wb["Summary"]
        dark = PatternFill(fill_type='solid', fgColor='bababa')
        light = fill=PatternFill(fill_type='solid', fgColor='dddddd')
        align = alignment=Alignment(horizontal="centerContinuous")

        border = Side(border_style="thin", color="000000")
        ws.sheet_view.showGridLines = False  # No grid lines
        ws.column_dimensions['B'].width = 30  # Change width of labels

        # Header
        ws['B2'] = "Financial Overview"
        ws['B2'].font = Font(bold=True)
        ws['B2'].fill = dark
        if self.is_unit == 'm':
            ws['B3'] = "($ in millions of U.S. Dollar)"
        else:
            ws['B3'] = "($ in billions of U.S. Dollar)"
        ws['B3'].font = Font(italic=True)
        ws['B3'].fill = dark

        # Summary Financials
        end = chr(ord('C') + len(years))
        ws.column_dimensions['C'].width = 15
        ws['C5'] = "Summary Financials"
        style_range(ws, 'C5', end + '5', fill=dark, font=Font(bold=True), alignment=align)
        ws['C6'] = "FYE " + self.fye
        style_range(ws, 'C6', end + '6', alignment=align,
                    border=Border(top=border, bottom=border))
        for i in range(len(years)):
            ws[chr(ord('D') + i) + '7'] = years[i]
        style_range(ws, 'D7', end + '7', font=Font(bold=True, underline="single"),
                    alignment=Alignment(horizontal="center"))
        ws['C8'], ws['C9'] = "Revenue", "Growth %"
        ws['C11'], ws['C12'], ws['C13'] = "Gross Profit", "Margin %", "Growth %"
        ws['C15'], ws['C16'], ws['C17'] = "EBITDA", "Margin %", "Growth %"
        ws['C19'],  ws['C20'], ws['C22'], ws['C23'] = "EPS", "Growth %", "ROA", "ROE"
        for i in range(len(years)):
            cur_col = chr(ord('D') + i)
            revenue = excel_cell(
                self.is_df, searched_label(self.is_df.index, "total sales"), years[i]
            )
            prev_revenue = chr(ord(revenue[0]) - 1) + revenue[1:]
            ws[cur_col + '8'] = "='{}'!{}".format(IS, revenue)
            ws[cur_col + '9'] = ws[cur_col + '8'].value + "/'{}'!{}-1".format(
                IS, prev_revenue
            )
            gross_profit = excel_cell(
                self.is_df, searched_label(self.is_df.index, "gross income"), years[i]
            )
            prev_gross_profit = chr(ord(gross_profit[0]) - 1) + gross_profit[1:]
            ws[cur_col + '11'] = "='{}'!{}".format(IS, gross_profit)
            ws[cur_col + '12'] = '=' + cur_col + '11/' + cur_col + '8'
            ws[cur_col + '13'] = ws[cur_col + '11'].value  + "/'{}'!{} - 1".format(
                IS, prev_gross_profit
            )
            ebitda = excel_cell(self.is_df, searched_label(self.is_df.index, "ebitda"), years[i])
            prev_ebitda = chr(ord(ebitda[0]) - 1) + ebitda[1:]
            ws[cur_col + '15'] = "='{}'!{}".format(IS, ebitda)
            ws[cur_col + '16'] = '=' + cur_col + '15/' + cur_col + '8'
            ws[cur_col + '17'] = ws[cur_col + '15'].value  + "/'{}'!{} - 1".format(
                IS, prev_ebitda
            )
            eps = excel_cell(self.is_df, searched_label(self.is_df.index, "eps diluted"), years[i])
            prev_eps = chr(ord(eps[0]) - 1) + eps[1:]
            ws[cur_col + '19'] = "='{}'!{}".format(IS, eps)
            ws[cur_col + '20'] = ws[cur_col + '19'].value  + "/'{}'!{} - 1".format(
                IS, prev_eps
            )
            net_income = excel_cell(
                self.is_df, searched_label(self.is_df.index, "net income"), years[i]
            )
            total_assets = excel_cell(
                self.bs_df, searched_label(self.bs_df.index, "total assets"), years[i]
            )
            total_equity = excel_cell(
                self.bs_df, searched_label(self.bs_df.index, "total equity"), years[i]
            )
            temp = "='{}'!{}/'{}'!".format(IS, net_income, BS)
            ws[cur_col + '22'] = temp + total_assets + self.extra_bs
            ws[cur_col + '23'] = temp + total_equity + self.extra_bs
        style_range(ws, 'C8', end + '8', font=Font(bold=True), currency=True)
        style_range(ws, 'C11', end + '11', font=Font(bold=True), currency=True)
        style_range(ws, 'C15', end + '15', font=Font(bold=True), currency=True)
        style_range(ws, 'C19', end + '19', font=Font(bold=True))
        style_range(ws, 'C9', end + '9', font=Font(italic=True), percentage=True)
        style_range(ws, 'C12', end + '12', font=Font(italic=True), percentage=True)
        style_range(ws, 'C13', end + '13', font=Font(italic=True), percentage=True)
        style_range(ws, 'C16', end + '16', font=Font(italic=True), percentage=True)
        style_range(ws, 'C17', end + '17', font=Font(italic=True), percentage=True)
        style_range(ws, 'C20', end + '20', font=Font(italic=True), percentage=True)
        style_range(ws, 'C22', end + '22', font=Font(italic=True), percentage=True)
        style_range(ws, 'C23', end + '23', font=Font(italic=True), percentage=True)
        style_range(ws, 'C24', end + '24', border=Border(top=border))
        style_range(ws, 'B6', 'B23', border=Border(right=border))
        style_range(
            ws, chr(ord(end) + 1) + '6', chr(ord(end) + 1) + '23', border=Border(left=border)
        )

        # Captilization
        start, end = chr(ord(end) + 3), chr(ord(end) + 4)
        ws.column_dimensions[start].width = 25
        ws[start + '5'], ws[start + '7'] = "Capitalization", "Cash and ST Investments"
        ws[start + '8'], ws[start + '9'], ws[start + '11'] = "Debt", "Net Debt", "Share Price"
        ws[start + '12'], ws[start + '13'] = "D/SO", "Market Cap"
        ws[start + '14'] = "Enterprise Value"
        style_range(ws, start + '5', end + '5', font=Font(bold=True), fill=dark,
                    alignment=align, border=Border(bottom=border))
        style_range(ws, start + '9', end + '9', font=Font(bold=True), border=Border(top=border))
        style_range(ws, start + '13', end + '13', border=Border(top=border))
        style_range(
            ws, start + '14', end + '14', font=Font(bold=True),
            border=Border(top=border, bottom=border)
        )
        style_range(
            ws, chr(ord(start) - 1) + '6', chr(ord(start) - 1) + '14', border=Border(right=border)
        )
        style_range(
            ws, chr(ord(end) + 1) + '6', chr(ord(end) + 1) + '14', border=Border(left=border)
        )
        ws[end + '9'] = '={}8-{}7'.format(end, end)
        ws[end + '12'] = "='{}'!{}".format(IS, excel_cell(
            self.is_df,
            searched_label(self.is_df.index, "diluted shares outstanding"),
            self.is_df.columns[-self.yrs_to_predict]
        ))
        ws[end + '13'] = '={}11*{}12'.format(end, end)
        ws[end + '14'] = '={}13+{}9'.format(end, end)
        share_price = start + '11'
        ev  = start + '14'
        d_so = start + '12'

        # Long term cash growth rate
        ws[start + '16'] = "Long Term Cash Growth Rate"
        ws[start + '17'], ws[start + '18'], ws[start + '19'] = "Bull", "Upside", "Base"
        ws[start + '20'], ws[start + '21'] = "Downside", "Bear"
        style_range(ws, start + '17', start + '21', light)
        lt_cash_rate = start + '16'

        # Valuation
        pred_years = [int(yr[:-1]) for yr in years if isinstance(yr, str)]
        num_pred = len(pred_years)
        start, end = chr(ord(end) + 3), chr(ord(end) + 2 + 3 * num_pred)
        ws[start + '5'] = "Valuation"
        style_range(ws, start + '5', end + '5', fill=dark, font=Font(bold=True),
                    border=Border(bottom=border), alignment=align)
        for i in range(num_pred):
            yr_start = chr(ord(start) + i * num_pred)
            yr_mid = chr(ord(start) + 1 + i * num_pred)
            yr_end = chr(ord(start) + 2 + i * num_pred)
            ws.column_dimensions[yr_start].width = 11
            ws.column_dimensions[yr_mid].width = 11
            ws.column_dimensions[yr_end].width = 11
            ws[yr_mid + '6'] = pred_years[i]
            ws[yr_mid + '6'].font = Font(bold=True)
            ws[yr_mid + '6'].alignment = Alignment(horizontal='center')
            ws[yr_start + '6'].border = Border(left=border)
            ws[yr_end + '6'].border = Border(right=border)
            ws[yr_start + '7'], ws[yr_mid + '7'] = 'P/E', 'EV/EBITDA'
            ws[yr_end + '7'] = 'EV/Sales'
            ws[yr_start + '8'] = "={}/'{}'!{}".format(share_price, IS,
                excel_cell(
                    self.is_df, searched_label(self.is_df.index, "eps diluted"),
                    str(pred_years[i]) + 'E'
                )
            )
            ws[yr_mid + '8'] = "={}/'{}'!{}".format(ev, IS,
                excel_cell(
                    self.is_df, searched_label(self.is_df.index, "ebitda"),
                    str(pred_years[i]) + 'E'
                )
            )
            ws[yr_end + '8'] = "={}/'{}'!{}".format(ev, IS,
                excel_cell(
                    self.is_df, searched_label(self.is_df.index, "total sales"),
                    str(pred_years[i]) + 'E'
                )
            )
        style_range(ws, start + '7', end + '7', font=Font(bold=True),
                    alignment=Alignment(horizontal='center'),
                    border=Border(top=border, left=border, bottom=border, right=border))
        style_range(ws, start + '8', end + '8', multiple=True,
                    border=Border(top=border, left=border, bottom=border, right=border))

        # Discounted Cash Flow
        end = chr(ord(start))
        start = chr(ord(start) - 1)
        ws.column_dimensions[start].width = 20
        ws[start + '11'] = "Discounted Cash Flow"
        style_range(ws, start + '11', end + '11', border=Border(bottom=border),
                    font=Font(bold=True), fill=light, alignment=align)
        ws[start + '12'], ws[start + '13'] = "Cost of Equity", "Terminal Value"
        ws[start + '14'], ws[start + '15'] = "Total Equity Value", "Target Price"
        ws[start + '15'].font = Font(bold=True)
        style_range(ws, start + '12', start + '15',
                    border=Border(bottom=border, top=border, left=border, right=border))
        style_range(ws, end + '12', end + '15',
                    border=Border(bottom=border, top=border, left=border, right=border))
        div_per_share = "'{}'!{}".format(IS, excel_cell(
            self.is_df, searched_label(self.is_df.index, "dividends per share"),
            self.is_df.columns[-self.yrs_to_predict]
        ))
        prev_div_per_share = "'{}'!{}".format(IS, excel_cell(
            self.is_df, searched_label(self.is_df.index, "dividends per share"),
            self.is_df.columns[-self.yrs_to_predict - 1]
        ))
        ws[end + '12'] = "={}/{}+{}/{}-1".format(div_per_share, share_price, div_per_share,
                                                 prev_div_per_share)
        ws[end + '13'] = "='{}'!{}*(1+{})/({}-{})".format(CF, excel_cell(
            self.cf_df, searched_label(self.cf_df.index, "levered free cash flow"),
            self.cf_df.columns[-1]
        ), lt_cash_rate, end + '12', lt_cash_rate)
        total_equity_val = '='
        for i, yr in enumerate(pred_years):
            total_equity_val += "'{}'!{}/(1+{})^{}+".format(CF, excel_cell(
                self.cf_df,
                searched_label(self.cf_df.index, "levered free cash flow"), str(yr) + 'E'
            ), end + '12', i + 1)
        total_equity_val += '{}/(1+{})^{}'.format(end + '13', end + '12', total_equity_val[-2])
        ws[end + '14'] = total_equity_val
        ws[end + '15'] = '={}/{}'.format(end + '14', d_so)

    def add_case_cells(self):
        ws = self.wb[IS]
        def style_cell(ws, excel_cell):
            ws[excel_cell].font = Font(bold=True, color='0000FF')
            ws[excel_cell].fill = PatternFill(fill_type='solid', fgColor='FFF2CC')
            border = Side(border_style="medium", color="000000")
            ws[excel_cell].border = Border(top=border, left=border, bottom=border, right=border)
            ws[excel_cell].alignment = Alignment(horizontal="center")
        ws[self.case_col + '2'] = 3
        style_cell(ws, self.case_col + '2')
        ws[chr(ord(self.case_col) - 1) + '2'] = "Case"
        ws[chr(ord(self.case_col) - 1) + '2'].font = Font(bold=True)
        ws[chr(ord(self.case_col) + 1) + '1'] = "1 - Bull"
        ws[chr(ord(self.case_col) + 1) + '2'] = "2 - Upside"
        ws[chr(ord(self.case_col) + 1) + '3'] = "3 - Base"
        ws[chr(ord(self.case_col) + 1) + '4'] = "4 - Downside"
        ws[chr(ord(self.case_col) + 1) + '5'] = "5 - Bear"

        rates = [0.01, -0.01, 0.005]
        for i, rate_cell in enumerate(self.case_rate_cells):
            ws[rate_cell] = rates[i]
            style_cell(ws, rate_cell)

    def save_wb(self):
        self.wb.save(
            "output/output_{}_{}.xlsx".format(self.name, datetime.now().strftime('%H-%M-%S'))
        )


def main():
    """Call all methods."""
    if os.path.exists('output'):
        rmtree('output')
    os.mkdir('output')

    growth_rates = [0.2, 0.2, 0.2, 0.2, 0.2]
    for i in NAME:
        print(i)
        vm = ValuationMachine(i, growth_rates)
        vm.read()
        vm.preprocess()
        vm.get_units()
        vm.slice_data()
        vm.process_is()
        vm.process_bs()
        vm.process_cf()
        vm.style()
        vm.add_summary()
        vm.add_case_cells()
        vm.save_wb()


if __name__ == "__main__":
    main()
