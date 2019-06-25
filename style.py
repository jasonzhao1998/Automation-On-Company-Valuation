"""Implementation for rendering excel output's style."""
import pandas as pd
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from helper import excel_cell, searched_label


def style_range(ws, start, end, fill=None, font=None, border=None,
                alignment=None, percentage=False, currency=False, multiple=False):
    """Change excel style for a column or row."""
    letter1, num1 = start[0], start[1:]
    letter2, num2 = end[0], end[1:]
    if num1 == num2:  # row
        for i in range(ord(letter2) - ord(letter1) + 1):
            if font:
                ws[chr(ord(letter1) + i) + num1].font = font
            if fill:
                ws[chr(ord(letter1) + i) + num1].fill = fill
            if border:
                ws[chr(ord(letter1) + i) + num1].border = border
            if alignment:
                ws[chr(ord(letter1) + i) + num1].alignment = alignment
            if currency:
                ws[chr(ord(letter1) + i) + num1].number_format = '$#,##'
            elif percentage:
                ws[chr(ord(letter1) + i) + num1].number_format = '0.00%'
            elif multiple:
                ws[chr(ord(letter1) + i) + num1].number_format = '0.0 x'
    elif letter1 == letter2:  # column
        for i in range(int(num1), int(num2) + 1):
            if font:
                ws[letter1 + str(i)].font = font
            if fill:
                ws[letter1 + str(i)].fill = fill
            if border:
                ws[letter1 + str(i)].border = border
            if alignment:
                ws[letter1 + str(i)].alignment = alignment
    else:
        print("ERROR: style_range", start, end)
        exit(1)


def style_ws(ws, sheet_name, is_df, bs_df, cf_df, fye, unit):
    """Change excel style for a worksheet."""
    if sheet_name == "Income Statement":
        cur_df = is_df
    elif sheet_name == "Balance Sheet":
        cur_df = bs_df
    elif sheet_name == "Cashflow Statement":
        cur_df = cf_df

    border = Side(border_style="thin", color="000000")

    # Insert empty column to beginning
    ws.insert_cols(1)

    letter, num = ws.dimensions.split(':')[1][0], ws.dimensions.split(':')[1][1:]

    ws.sheet_view.showGridLines = False  # No grid lines
    ws.move_range('C1:' + letter + '1', rows=4)  # Move year row down
    ws.column_dimensions['B'].width = 50  # Change width of labels

    ws['B2'] = sheet_name
    ws['B2'].font = Font(bold=True)
    ws['B2'].fill = PatternFill(fill_type='solid', fgColor='bababa')
    if unit == 'm':
        ws['B3'] = "($ in millions of U.S. Dollar)"
    else:
        ws['B3'] = "($ in billions of U.S. Dollar)"
    ws['B3'].font = Font(italic=True)
    style_range(ws, 'B3', letter + '3', fill=PatternFill(fill_type='solid', fgColor='bababa'))

    # Central element Annual
    ws[chr((ord('C') + ord(letter)) // 2) + '3'] = "Annual"
    ws[chr((ord('C') + ord(letter)) // 2) + '3'].font = Font(bold=True)
    ws[chr((ord('C') + ord(letter)) // 2) + '4'] = "FYE " + fye
    # Center across selection
    temp = Alignment(horizontal='centerContinuous')
    ws[chr((ord('C') + ord(letter)) // 2) + '3'].alignment = temp
    ws[chr((ord('C') + ord(letter) + 1) // 2) + '3'].alignment = temp
    ws[chr((ord('C') + ord(letter)) // 2) + '4'].alignment = temp
    ws[chr((ord('C') + ord(letter) + 1) // 2) + '4'].alignment = temp

    # Year row style
    style_range(ws, 'C5', letter + '5', font=Font(bold=True, underline="single"),
                border=Border(top=border, bottom=border),
                alignment=Alignment(horizontal="center", vertical="center"))

    # Label column
    style_range(ws, 'B7', 'B' + num, fill=PatternFill(fill_type='solid', fgColor='dddddd'))

    # Style sum rows
    for cell in [letter + str(i + 1) for i in range(int(num) - 1)]:
        if isinstance(ws[cell].value, str) and 'SUM' in ws[cell].value and len(ws[cell].value) < 30:
            num = cell[1:]
            ws['B' + num].font = Font(bold=True)
            style_range(ws, 'C' + num, letter + num, font=Font(bold=True),
                        border=Border(top=border), currency=True)

    # Style specific rows
    def style_row(ws, label, df, border_bool=True, bold_bool=True, italic_bool=False,
                  underline=None, currency=False):
        num = str(int(excel_cell(df, searched_label(df.index, label), df.columns[0])[1:]))
        ws['B' + num].font = Font(bold=True, underline=underline)
        border_style = Border(top=border) if border_bool else Border()
        if bold_bool:
            font_style = Font(bold=True)
        elif italic_bool:
            font_style = Font(italic=True)
        else:
            font_style = Font()
        style_range(ws, 'C' + num, letter + num, font=font_style, border=border_style,
                    currency=currency)

    if sheet_name == "Income Statement":
        style_row(ws, "total sales", cur_df, False, currency=True)
        style_row(ws, "gross income", cur_df, currency=True)
        style_row(ws, "ebit operating income", cur_df, currency=True)
        style_row(ws, "pretax income", cur_df, currency=True)
        style_row(ws, "net income", cur_df, currency=True)
    elif sheet_name == "Balance Sheet":
        style_row(ws, "total shareholder equity", cur_df, bold_bool=False, border_bool=False,
                  currency=True)
        style_row(ws, "total liabilit shareholder equity", cur_df, border_bool=False)
        style_row(ws, "balance", cur_df, border_bool=False, bold_bool=False, italic_bool=True)
    elif sheet_name == "Cashflow Statement":
        style_row(ws, "net operating cash flow cf", cur_df, currency=True)
        style_row(ws, "cash balance", cur_df, border_bool=False, currency=True)
    style_row(ws, "driver ratio", cur_df, underline="single", border_bool=False)
    driver_df = cur_df.loc["Driver Ratios":]

    # Driver ratios style
    driver_i = cur_df.index.get_loc("Driver Ratios")
    for i, ratio in enumerate(driver_df.iloc[1:].index):
        if pd.isna(ratio) or ratio == "DPO" or ratio == "DSO" or ratio == "Levered Free Cash Flow":
            continue
        start = 'C' + str(driver_i + i + 4)
        end = letter + str(int(start[1:]))
        if ratio == "Bull" or ratio == "Upside" or ratio == "Base" or ratio == "Downside" or ratio == "Bear":
            style_range(ws, start, end, font=Font(italic=True))
        style_range(ws, start, end, percentage=True)
